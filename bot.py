# ============================================================
# BOT DE MESA DE AYUDA TECNICA - Simulador en consola
# Trabajo Practico Integrador - Organizacion Empresarial
# UTN - Tecnicatura Universitaria en Programacion
# ============================================================

import openpyxl
from datetime import date

# ------------------------------------------------------------
# Ruta al archivo de Excel
# ------------------------------------------------------------
RUTA_EXCEL = "data/base_datos_mesa_ayuda.xlsx"

# ------------------------------------------------------------
# Estados Operativos del Bot
# ------------------------------------------------------------
ESPERANDO_LEGAJO     = "ESPERANDO_LEGAJO"
ESPERANDO_CATEGORIA  = "ESPERANDO_CATEGORIA"
ESPERANDO_DESCRIPCION= "ESPERANDO_DESCRIPCION"
BUSCANDO_SOLUCION    = "BUSCANDO_SOLUCION"
ESPERANDO_CONFIRM    = "ESPERANDO_CONFIRMACION"
CREANDO_TICKET       = "CREANDO_TICKET"
RESUELTO             = "RESUELTO"
DERIVADO             = "DERIVADO"

# ------------------------------------------------------------
# Funciones de Base de Datos
# ------------------------------------------------------------

def buscar_usuario(legajo):
    #Busca un usuario por legajo en la hoja Usuarios.
    try:
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        hoja = wb["Usuarios"]
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if str(fila[0]) == str(legajo):
                return {"legajo": fila[0], "nombre": fila[1],
                        "area": fila[2], "email": fila[3]}
        return None
    except FileNotFoundError:
        print("[ERROR] No se encontro el archivo Excel.")
        return None


def buscar_solucion(categoria, descripcion):
    #Busca en la FAQ si hay una solucion para el problema.
    try:
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        hoja = wb["BaseConocimiento"]
        descripcion_lower = descripcion.lower()
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            cat_fila    = str(fila[1]).lower()
            palabras    = str(fila[2]).lower().split(",")
            solucion    = fila[3]
            if cat_fila == categoria.lower():
                for palabra in palabras:
                    if palabra.strip() in descripcion_lower:
                        return solucion
        return None
    except FileNotFoundError:
        print("[ERROR] No se encontro el archivo Excel.")
        return None


def crear_ticket(usuario, categoria, descripcion, prioridad):
    #Registra un nuevo ticket en la hoja Tickets del Excel.
    try:
        wb   = openpyxl.load_workbook(RUTA_EXCEL)
        hoja = wb["Tickets"]
        ultima_fila = hoja.max_row
        nuevo_id    = "T-" + str(1000 + ultima_fila)
        hoja.append([
            nuevo_id,
            str(date.today()),
            usuario["legajo"],
            usuario["nombre"],
            usuario["area"],
            categoria,
            descripcion,
            prioridad,
            "Derivado N2",
            ""
        ])
        wb.save(RUTA_EXCEL)
        return nuevo_id
    except FileNotFoundError:
        print("[ERROR] No se encontro el archivo Excel.")
        return "T-ERROR"


def determinar_prioridad(categoria):
    #Asigna prioridad segun la categoria del problema.
    if categoria == "Hardware":
        return "Alta"
    elif categoria == "Red":
        return "Media"
    else:
        return "Baja"


# ------------------------------------------------------------
# Funcion principal del bot
# ------------------------------------------------------------

def ejecutar_bot():

    print("=" * 55)
    print("  SIMULADOR - BOT DE MESA DE AYUDA TECNICA")
    print("  Trabajo Practico Integrador | OE | UTN TUP")
    print("=" * 55)
    print()

    # Variables de sesion
    estado    = ESPERANDO_LEGAJO
    usuario   = None
    categoria = None
    descripcion = None

    # --------------------------------------------------------
    # Estado Inicial:
    # --------------------------------------------------------
    print("Bot: Hola! Soy el asistente de Soporte IT.")
    print("     Para comenzar, ingresa tu numero de legajo.")
    print("    (Solo numeros, ejemplo: 4521 | o escribi salir para cancelar)")
    print()

    while True:
        entrada = input("Vos: ").strip()

        if entrada.lower() == "salir":
            print()
            print("Bot: Sesion cancelada. Hasta luego!")
            break

        # ----------------------------------------------------
        # ESPERANDO_LEGAJO  ESPERANDO_CATEGORIA o ERROR
        # ----------------------------------------------------
        if estado == ESPERANDO_LEGAJO:
            if entrada.isdigit():
                usuario = buscar_usuario(entrada)
                if usuario:
                    estado   = ESPERANDO_CATEGORIA
                    print()
                    print(f"Bot: Bienvenido/a, {usuario['nombre']} ({usuario['area']})!")
                    print("     Selecciona la categoria del problema:")
                    print("     1) Hardware")
                    print("     2) Software")
                    print("     3) Red")
                    print()
                else:
                    print()
                    print("Bot: [ERROR] No encontre ese legajo en el sistema.")
                    print("     Por favor, ingresa un legajo valido.")
                    print()
            else:
                print()
                print("Bot: [ERROR] El legajo debe contener solo numeros.")
                print("     Por favor ingresa solo digitos. Ejemplo: 4521")
                print()

        # ----------------------------------------------------
        # ESPERANDO_CATEGORIA ESPERANDO_DESCRIPCION o ERROR
        # ----------------------------------------------------
        elif estado == ESPERANDO_CATEGORIA:
            categorias = {"1": "Hardware", "2": "Software", "3": "Red"}
            if entrada in categorias:
                categoria = categorias[entrada]
                estado    = ESPERANDO_DESCRIPCION
                print()
                print(f"Bot: Categoria seleccionada: {categoria}.")
                print("      Describe brevemente el problema.")
                print()
            else:
                print()
                print("Bot: [ERROR] Opcion no valida.")
                print("     Escribi 1 (Hardware), 2 (Software) o 3 (Red).")
                print()

        # ----------------------------------------------------
        # ESPERANDO_DESCRIPCION  BUSCANDO_SOLUCION o ERROR
        # ----------------------------------------------------
        elif estado == ESPERANDO_DESCRIPCION:
            if len(entrada) >= 3:
                descripcion = entrada
                estado = BUSCANDO_SOLUCION
                print()
                print("Bot: Buscando una solucion en nuestra base de conocimiento...")
                print()

                solucion = buscar_solucion(categoria, descripcion)

                if solucion:
                    # GATEWAY 1: Problema conocido en FAQ  SI 
                    estado   = ESPERANDO_CONFIRM
                    print(f"Bot: Encontre una solucion!")
                    print(f"     >> {solucion}")
                    print()
                    print("Bot: Segui estos pasos e indicame:")
                    print("     Se resolvio el problema? (si / no)")
                    print()
                else:
                    # GATEWAY 1: Problema conocido en FAQ = NO
                    estado = CREANDO_TICKET
                    prioridad  = determinar_prioridad(categoria)
                    id_ticket  = crear_ticket(usuario, categoria,
                                              descripcion, prioridad)
                    print(f"Bot: No encontre una solucion automatica para este problema.")
                    print(f"     Se genero el ticket {id_ticket}.")
                    print(f"     Prioridad: {prioridad}.")
                    print("     Un tecnico de Nivel 2 se comunicara en las proximas 4 hs.")
                    print()
                    estado = DERIVADO
                    print("Bot: Tu caso quedo registrado. Que tengas un buen dia!")
                    print()
                    break

            else:
                print()
                print("Bot: [ERROR] La descripcion es muy corta.")
                print("     Por favor describe el problema con mas detalle (minimo 3 caracteres).")
                print()

        # ----------------------------------------------------
        # ESPERANDO_CONFIRMACION  RESUELTO o CREANDO_TICKET
        # ----------------------------------------------------
        elif estado == ESPERANDO_CONFIRM:
            respuesta = entrada.lower()
            if respuesta in ["si", "sí", "s"]:
                # GATEWAY 2: Se resolvio = SI 
                estado = RESUELTO
                print()
                print("Bot: Excelente! Me alegra que se haya resuelto.")
                print("     El caso se cierra como RESUELTO.")
                print("     Que tengas un buen dia!")
                print()
                break

            elif respuesta in ["no", "n"]:
                # GATEWAY 2: Se resolvio = NO 
                estado    = CREANDO_TICKET
                prioridad = determinar_prioridad(categoria)
                id_ticket = crear_ticket(usuario, categoria,
                                         descripcion, prioridad)
                print()
                print("Bot: Entendido. Escalo el caso a un especialista.")
                print(f"     Ticket generado: {id_ticket}.")
                print(f"     Prioridad: {prioridad}.")
                print("      Te contactaran en las proximas 4 horas.")
                print()
                estado = DERIVADO
                print("Bot: Que tengas un buen dia!")
                print()
                break

            else:
                print()
                print("Bot: [ERROR] No entendi tu respuesta.")
                print("     Por favor, respondé solamente 'si' o 'no'.")
                print()


# ------------------------------------------------------------
# Entrada
# ------------------------------------------------------------
if __name__ == "__main__":
    ejecutar_bot()
